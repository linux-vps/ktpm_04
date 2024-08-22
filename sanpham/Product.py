from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
import openpyxl

# Định nghĩa trình duyệt
def open_edge():
    options = Options()
    options.add_argument('--start-maximized')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    service = Service(executable_path="msedgedriver.exe")
    driver = webdriver.Edge(service=service, options=options)
    return driver

# Đọc dữ liệu từ file excel
def read_data_from_excel(file_path, min_row, max_row):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row):
        row_data = [
            row[4].value,
            row[5].value,
            row[6].value,
            row[7].value,
            row[8].value,
            row[9].value,
            row[10].value
        ]
        data.append(row_data)
    return data

# Ghi dữ liệu ra file excel
def write_data_to_excel(file_path, min_row, max_row, resutl_column, result):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    index = 0
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row):
        row[resutl_column].value = result[index]
        index += 1
    workbook.save(file_path)

# Điều hướng đến trang đăng nhập
def go_to_login_page(driver):
    try:
        # Đợi nút đăng nhập xuất hiện và nhấp vào nó
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#navbarCollapse > div.d-flex.me-0 > a"))).click()
        # Xác nhận chuyển hướng đến trang đăng nhập
        WebDriverWait(driver, 10).until(EC.url_contains("/login"))
        time.sleep(1)
    except Exception as e:
        print(f"An error occurred: {e}")

# Đăng nhập vào tài khoản người dùng
def login(driver, username, password):
    # Giả sử có form đăng nhập trên trang
    driver.find_element(By.CSS_SELECTOR, "#layoutAuthentication_content > main > div > div > div > div > div.card-body > form > div:nth-child(1) > input").send_keys(username)
    driver.find_element(By.CSS_SELECTOR, "#password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR, "#layoutAuthentication_content > main > div > div > div > div > div.card-body > form > div.mt-4.mb-0 > div > button").click()
    time.sleep(1)

# Chuyển hướng đến trang sản phẩm
def go_to_product_management(driver):
    try:
        driver.find_element(By.LINK_TEXT, "Sản phẩm").click()
    except Exception as e:
        print(f"An error occurred: {e}")

# Chuyển hướng đến trang thêm sản phẩm
def go_to_add_product(driver):
    try:
        driver.find_element(By.LINK_TEXT, "Thêm sản phẩm").click()
    except Exception as e:
        print(f"An error occurred: {e}")

# Kiểm tra có xuất hiện lỗi hay không
def check(driver):
    time.sleep(1)
    errMsg = driver.find_elements(By.CLASS_NAME, "invalid-feedback")
    if errMsg:
        return False
    else:
        return True

# Thêm mới sản phẩm
def add_product(driver, name, price, detailDesc, shortDesc, quantity, factory, target):
    go_to_product_management(driver)
    go_to_add_product(driver)

    name_field = driver.find_element(By.ID, "name")
    price_field = driver.find_element(By.ID, "price")
    detailDesc_field = driver.find_element(By.ID, "detailDesc")
    shortDesc_field = driver.find_element(By.ID, "shortDesc")
    quantity_field = driver.find_element(By.ID, "quantity")
    factory_field = driver.find_element(By.ID, "factory")
    target_field = driver.find_element(By.ID, "target")

    name_field.clear()
    price_field.clear()
    detailDesc_field.clear()
    shortDesc_field.clear()
    quantity_field.clear()

    if name != None:
        name_field.send_keys(name)
    if price != None:
        price_field.send_keys(price)
    if detailDesc != None:
        detailDesc_field.send_keys(detailDesc)
    if shortDesc != None:
        shortDesc_field.send_keys(shortDesc)
    if quantity != None:
        quantity_field.send_keys(quantity)
    if factory != None:
        factory_field.send_keys(factory)
    if target != None:
        target_field.send_keys(target)

    driver.find_element(By.CSS_SELECTOR, ".btn-primary").click()
    add_successful = check(driver)
    if add_successful:
        return "Thành công"
    else:
        time.sleep(1)
        errMsg = driver.find_element(By.CLASS_NAME, "invalid-feedback").text
        return errMsg

# Quản lý sản phẩm
def product_management(username, password, list_add):
    driver = open_edge();
    driver.get("http://localhost:8080/")
    go_to_login_page(driver)
    login(driver, username, password)
    result = []
    for name, price, detailDesc, shortDesc, quantity, factory, target in list_add:
        result.append(add_product(driver, name, price, detailDesc, shortDesc, quantity, factory, target))
    print("Đã xong")
    driver.quit()
    return result

if __name__ == '__main__':
    username = "manager@gmail.com"
    password = "123456"
    file_path = "TestPlan_Product.xlsx"
    min_row = 3
    max_row = 12
    result_column = 12
    list_add = read_data_from_excel(file_path, min_row, max_row)
    result = product_management(username, password, list_add)
    write_data_to_excel(file_path, min_row, max_row, result_column, result)



